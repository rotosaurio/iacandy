"""
Generador de reportes con exportación a Excel para grandes volúmenes.

Este módulo maneja la exportación de resultados de consultas SQL a diferentes
formatos, con soporte especial para streaming de grandes datasets.
"""

import os
import time
import threading
from typing import Dict, List, Optional, Any, Iterator, Callable
from datetime import datetime
from pathlib import Path
from dataclasses import dataclass

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.chart import BarChart, LineChart, PieChart, Reference
import xlsxwriter

from config import config, StatusMessages
from database import db, QueryResult
from utils import logger, timing_decorator, DataFormatter, DataAnalyzer


@dataclass
class ExportProgress:
    """Estado del progreso de exportación."""
    total_rows: int
    processed_rows: int
    current_batch: int
    status: str
    error: Optional[str] = None
    file_path: Optional[str] = None
    start_time: datetime = None
    
    def __post_init__(self):
        if self.start_time is None:
            self.start_time = datetime.now()
    
    @property
    def progress_percentage(self) -> float:
        if self.total_rows == 0:
            return 100.0
        return min((self.processed_rows / self.total_rows) * 100, 100.0)
    
    @property
    def elapsed_time(self) -> float:
        return (datetime.now() - self.start_time).total_seconds()
    
    @property
    def estimated_remaining_time(self) -> float:
        if self.processed_rows == 0:
            return 0.0
        
        rate = self.processed_rows / self.elapsed_time
        remaining_rows = self.total_rows - self.processed_rows
        
        if rate > 0:
            return remaining_rows / rate
        return 0.0


class ExcelStyler:
    """Aplicador de estilos para archivos Excel."""
    
    @staticmethod
    def get_header_style() -> Dict[str, Any]:
        """Obtener estilo para encabezados."""
        return {
            'font': Font(bold=True, color='FFFFFF'),
            'fill': PatternFill(start_color='366092', end_color='366092', fill_type='solid'),
            'alignment': Alignment(horizontal='center', vertical='center'),
            'border': Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
        }
    
    @staticmethod
    def get_data_style() -> Dict[str, Any]:
        """Obtener estilo para datos."""
        return {
            'alignment': Alignment(horizontal='left', vertical='center'),
            'border': Border(
                left=Side(style='thin', color='CCCCCC'),
                right=Side(style='thin', color='CCCCCC'),
                top=Side(style='thin', color='CCCCCC'),
                bottom=Side(style='thin', color='CCCCCC')
            )
        }
    
    @staticmethod
    def get_number_style() -> Dict[str, Any]:
        """Obtener estilo para números."""
        style = ExcelStyler.get_data_style()
        style['alignment'] = Alignment(horizontal='right', vertical='center')
        style['number_format'] = '#,##0.00'
        return style
    
    @staticmethod
    def apply_styles(worksheet, start_row: int, end_row: int, columns: List[str]):
        """Aplicar estilos a un rango de celdas."""
        header_style = ExcelStyler.get_header_style()
        data_style = ExcelStyler.get_data_style()
        number_style = ExcelStyler.get_number_style()
        
        # Aplicar estilo a encabezados
        for col_num, col_name in enumerate(columns, 1):
            cell = worksheet.cell(row=1, column=col_num)
            cell.font = header_style['font']
            cell.fill = header_style['fill']
            cell.alignment = header_style['alignment']
            cell.border = header_style['border']
            
            # Auto-ajustar ancho de columna
            column_letter = worksheet.cell(row=1, column=col_num).column_letter
            worksheet.column_dimensions[column_letter].width = min(max(len(col_name) + 2, 10), 50)
        
        # Aplicar estilo a datos
        for row in range(start_row, end_row + 1):
            for col in range(1, len(columns) + 1):
                cell = worksheet.cell(row=row, column=col)
                
                # Determinar si es número
                if isinstance(cell.value, (int, float)):
                    cell.alignment = number_style['alignment']
                    cell.number_format = number_style['number_format']
                else:
                    cell.alignment = data_style['alignment']
                
                cell.border = data_style['border']


class ChartGenerator:
    """Generador de gráficos para Excel."""
    
    @staticmethod
    def create_summary_chart(worksheet, data_range: str, chart_type: str = 'bar') -> Any:
        """Crear gráfico de resumen."""
        if chart_type == 'bar':
            chart = BarChart()
            chart.title = "Resumen de Datos"
            chart.y_axis.title = 'Valores'
            chart.x_axis.title = 'Categorías'
        
        elif chart_type == 'line':
            chart = LineChart()
            chart.title = "Tendencia de Datos"
            chart.y_axis.title = 'Valores'
            chart.x_axis.title = 'Tiempo'
        
        elif chart_type == 'pie':
            chart = PieChart()
            chart.title = "Distribución de Datos"
        
        else:
            return None
        
        # Configurar datos del gráfico
        data = Reference(worksheet, range_string=data_range)
        chart.add_data(data, titles_from_data=True)
        
        return chart


class StreamingExporter:
    """Exportador con streaming para grandes volúmenes."""
    
    def __init__(self, progress_callback: Callable[[ExportProgress], None] = None):
        self.progress_callback = progress_callback
        self._cancel_export = False
    
    def cancel_export(self):
        """Cancelar exportación en curso."""
        self._cancel_export = True
    
    @timing_decorator("Streaming Excel Export")
    def export_to_excel_streaming(self, sql: str, file_path: str, 
                                 sheet_name: str = "Datos") -> ExportProgress:
        """Exportar query grande a Excel usando streaming."""
        progress = ExportProgress(
            total_rows=0,
            processed_rows=0,
            current_batch=0,
            status="Iniciando..."
        )
        
        try:
            # Crear directorio si no existe
            os.makedirs(os.path.dirname(file_path), exist_ok=True)
            
            # Obtener metadatos de la consulta primero
            preview_result = db.execute_query_limited(sql, limit=1)
            if preview_result.error:
                progress.status = "Error"
                progress.error = preview_result.error
                return progress
            
            columns = preview_result.columns
            
            # Estimar número total de filas (aproximado)
            count_sql = f"SELECT COUNT(*) FROM ({sql}) AS count_query"
            try:
                count_result = db.execute_query_limited(count_sql, limit=1)
                if not count_result.error and count_result.preview_data:
                    progress.total_rows = count_result.preview_data[0][0]
                else:
                    # Fallback: usar estimación
                    progress.total_rows = 10000  # Estimación conservadora
            except:
                progress.total_rows = 10000
            
            progress.status = "Creando archivo Excel..."
            self._update_progress(progress)
            
            # Usar XlsxWriter para mejor rendimiento con streaming
            workbook = xlsxwriter.Workbook(file_path, {'constant_memory': True})
            worksheet = workbook.add_worksheet(sheet_name)
            
            # Definir formatos
            header_format = workbook.add_format({
                'bold': True,
                'bg_color': '#366092',
                'font_color': 'white',
                'border': 1,
                'align': 'center',
                'valign': 'vcenter'
            })
            
            data_format = workbook.add_format({
                'border': 1,
                'align': 'left',
                'valign': 'vcenter'
            })
            
            number_format = workbook.add_format({
                'border': 1,
                'align': 'right',
                'valign': 'vcenter',
                'num_format': '#,##0.00'
            })
            
            # Escribir encabezados
            for col_num, column_name in enumerate(columns):
                worksheet.write(0, col_num, column_name, header_format)
                # Auto-ajustar ancho
                worksheet.set_column(col_num, col_num, min(max(len(column_name) + 2, 10), 50))
            
            progress.status = "Exportando datos..."
            self._update_progress(progress)
            
            # Exportar datos en batches
            current_row = 1
            batch_num = 0
            
            for batch in db.execute_query_streaming(sql):
                if self._cancel_export:
                    progress.status = "Cancelado"
                    workbook.close()
                    if os.path.exists(file_path):
                        os.remove(file_path)
                    return progress
                
                batch_num += 1
                progress.current_batch = batch_num
                
                # Escribir batch de datos
                for row_data in batch:
                    for col_num, value in enumerate(row_data):
                        if value is None:
                            worksheet.write(current_row, col_num, "", data_format)
                        elif isinstance(value, (int, float)):
                            worksheet.write(current_row, col_num, value, number_format)
                        elif isinstance(value, datetime):
                            worksheet.write(current_row, col_num, value.strftime('%Y-%m-%d %H:%M:%S'), data_format)
                        else:
                            worksheet.write(current_row, col_num, str(value), data_format)
                    
                    current_row += 1
                
                progress.processed_rows = current_row - 1
                progress.status = f"Procesando... {DataFormatter.format_number(progress.processed_rows)} registros"
                self._update_progress(progress)
                
                # Actualizar total si es necesario
                if progress.processed_rows > progress.total_rows:
                    progress.total_rows = progress.processed_rows + 1000
            
            # Finalizar
            progress.total_rows = progress.processed_rows
            workbook.close()
            
            progress.status = "Completado"
            progress.file_path = file_path
            self._update_progress(progress)
            
            logger.info(f"Exportación completada: {progress.processed_rows:,} registros en {file_path}")
            
            return progress
            
        except Exception as e:
            progress.status = "Error"
            progress.error = str(e)
            logger.error("Error en exportación streaming", e)
            return progress
    
    def export_to_csv_streaming(self, sql: str, file_path: str) -> ExportProgress:
        """Exportar query grande a CSV usando streaming."""
        progress = ExportProgress(
            total_rows=0,
            processed_rows=0,
            current_batch=0,
            status="Iniciando exportación CSV..."
        )
        
        try:
            import csv
            
            # Crear directorio si no existe
            os.makedirs(os.path.dirname(file_path), exist_ok=True)
            
            # Obtener columnas
            preview_result = db.execute_query_limited(sql, limit=1)
            if preview_result.error:
                progress.status = "Error"
                progress.error = preview_result.error
                return progress
            
            columns = preview_result.columns
            
            progress.status = "Exportando a CSV..."
            self._update_progress(progress)
            
            with open(file_path, 'w', newline='', encoding='utf-8-sig') as csvfile:
                writer = csv.writer(csvfile)
                
                # Escribir encabezados
                writer.writerow(columns)
                
                # Exportar datos en batches
                current_row = 0
                batch_num = 0
                
                for batch in db.execute_query_streaming(sql):
                    if self._cancel_export:
                        progress.status = "Cancelado"
                        return progress
                    
                    batch_num += 1
                    progress.current_batch = batch_num
                    
                    for row_data in batch:
                        # Formatear datos para CSV
                        formatted_row = []
                        for value in row_data:
                            if value is None:
                                formatted_row.append("")
                            elif isinstance(value, datetime):
                                formatted_row.append(value.strftime('%Y-%m-%d %H:%M:%S'))
                            else:
                                formatted_row.append(str(value))
                        
                        writer.writerow(formatted_row)
                        current_row += 1
                    
                    progress.processed_rows = current_row
                    progress.status = f"Exportando CSV... {DataFormatter.format_number(current_row)} registros"
                    self._update_progress(progress)
            
            progress.total_rows = progress.processed_rows
            progress.status = "Completado"
            progress.file_path = file_path
            self._update_progress(progress)
            
            logger.info(f"Exportación CSV completada: {progress.processed_rows:,} registros")
            return progress
            
        except Exception as e:
            progress.status = "Error"
            progress.error = str(e)
            logger.error("Error en exportación CSV", e)
            return progress
    
    def _update_progress(self, progress: ExportProgress):
        """Actualizar progreso."""
        if self.progress_callback:
            self.progress_callback(progress)


class ReportGenerator:
    """Generador principal de reportes."""
    
    def __init__(self):
        self.active_exports = {}
        self._export_counter = 0
    
    def create_enhanced_excel_report(self, query_result: QueryResult, 
                                   output_path: str = None,
                                   include_analysis: bool = True,
                                   include_charts: bool = True) -> str:
        """Crear reporte Excel enriquecido con análisis y gráficos."""
        try:
            if output_path is None:
                timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                output_path = os.path.join(config.export.output_directory, f"reporte_{timestamp}.xlsx")
            
            # Crear directorio si no existe
            os.makedirs(os.path.dirname(output_path), exist_ok=True)
            
            # Convertir datos a DataFrame
            if not query_result.preview_data:
                raise ValueError("No hay datos para exportar")
            
            df = pd.DataFrame(query_result.preview_data, columns=query_result.columns)
            
            with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
                # Hoja principal con datos
                df.to_excel(writer, sheet_name='Datos', index=False)
                
                # Aplicar estilos
                workbook = writer.book
                worksheet = writer.sheets['Datos']
                ExcelStyler.apply_styles(worksheet, 2, len(df) + 1, query_result.columns)
                
                # Hoja de análisis si se solicita
                if include_analysis:
                    analysis_data = DataAnalyzer.analyze_dataframe(df)
                    self._create_analysis_sheet(writer, analysis_data, query_result.sql)
                
                # Gráficos si se solicita
                if include_charts and len(df) > 0:
                    self._create_charts_sheet(writer, df)
                
                # Hoja de metadatos
                self._create_metadata_sheet(writer, query_result)
            
            logger.info(f"Reporte Excel creado: {output_path}")
            return output_path
            
        except Exception as e:
            logger.error("Error creando reporte Excel enriquecido", e)
            raise
    
    def start_streaming_export(self, sql: str, export_format: str = 'xlsx',
                             progress_callback: Callable[[ExportProgress], None] = None) -> str:
        """Iniciar exportación streaming en background."""
        self._export_counter += 1
        export_id = f"export_{self._export_counter}_{int(time.time())}"
        
        # Generar nombre de archivo
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        if export_format.lower() == 'xlsx':
            filename = f"export_{timestamp}.xlsx"
        elif export_format.lower() == 'csv':
            filename = f"export_{timestamp}.csv"
        else:
            raise ValueError(f"Formato no soportado: {export_format}")
        
        file_path = os.path.join(config.export.output_directory, filename)
        
        # Crear exportador
        exporter = StreamingExporter(progress_callback)
        
        # Función de exportación
        def export_worker():
            try:
                if export_format.lower() == 'xlsx':
                    progress = exporter.export_to_excel_streaming(sql, file_path)
                elif export_format.lower() == 'csv':
                    progress = exporter.export_to_csv_streaming(sql, file_path)
                
                # Mantener referencia del progreso
                self.active_exports[export_id] = progress
                
            except Exception as e:
                logger.error(f"Error en exportación {export_id}", e)
                progress = ExportProgress(0, 0, 0, "Error", str(e))
                self.active_exports[export_id] = progress
        
        # Iniciar en thread separado
        export_thread = threading.Thread(target=export_worker, daemon=True)
        export_thread.start()
        
        # Mantener referencia del exportador para cancelación
        self.active_exports[export_id] = exporter
        
        logger.info(f"Exportación streaming iniciada: {export_id}")
        return export_id
    
    def get_export_progress(self, export_id: str) -> Optional[ExportProgress]:
        """Obtener progreso de exportación."""
        return self.active_exports.get(export_id)
    
    def cancel_export(self, export_id: str) -> bool:
        """Cancelar exportación."""
        if export_id in self.active_exports:
            exporter = self.active_exports[export_id]
            if isinstance(exporter, StreamingExporter):
                exporter.cancel_export()
                return True
        return False
    
    def _create_analysis_sheet(self, writer, analysis_data: Dict[str, Any], sql: str):
        """Crear hoja de análisis."""
        analysis_df = pd.DataFrame([
            ['Consulta SQL', sql],
            ['Total de filas', analysis_data.get('rows', 0)],
            ['Total de columnas', analysis_data.get('columns', 0)],
            ['Completitud de datos', analysis_data.get('summary', {}).get('completeness', 'N/A')],
            ['Registros duplicados', analysis_data.get('summary', {}).get('duplicates', 0)],
            ['Columnas numéricas', analysis_data.get('summary', {}).get('numeric_columns', 0)],
            ['Columnas de texto', analysis_data.get('summary', {}).get('text_columns', 0)]
        ], columns=['Métrica', 'Valor'])
        
        analysis_df.to_excel(writer, sheet_name='Análisis', index=False)
        
        # Información de columnas
        if 'column_info' in analysis_data:
            col_analysis = []
            for col_name, col_info in analysis_data['column_info'].items():
                col_analysis.append([
                    col_name,
                    col_info.get('type', 'Unknown'),
                    col_info.get('non_null_count', 0),
                    col_info.get('unique_values', 0)
                ])
            
            col_df = pd.DataFrame(col_analysis, 
                                columns=['Columna', 'Tipo', 'Valores No Nulos', 'Valores Únicos'])
            
            # Agregar a la misma hoja con espacio
            start_row = len(analysis_df) + 3
            col_df.to_excel(writer, sheet_name='Análisis', startrow=start_row, index=False)
    
    def _create_charts_sheet(self, writer, df: pd.DataFrame):
        """Crear hoja con gráficos."""
        # Esta función requeriría más lógica específica para crear gráficos
        # Por ahora, creamos una hoja de resumen
        
        numeric_columns = df.select_dtypes(include=['number']).columns.tolist()
        
        if numeric_columns:
            summary_df = df[numeric_columns].describe()
            summary_df.to_excel(writer, sheet_name='Gráficos', index=True)
    
    def _create_metadata_sheet(self, writer, query_result: QueryResult):
        """Crear hoja con metadatos."""
        metadata_df = pd.DataFrame([
            ['Consulta SQL', query_result.sql],
            ['Tiempo de ejecución (segundos)', query_result.execution_time],
            ['Registros retornados', query_result.row_count],
            ['Hay más datos', 'Sí' if query_result.has_more_data else 'No'],
            ['Fecha de generación', datetime.now().strftime('%Y-%m-%d %H:%M:%S')],
            ['Columnas', ', '.join(query_result.columns)]
        ], columns=['Metadato', 'Valor'])
        
        metadata_df.to_excel(writer, sheet_name='Metadatos', index=False)
    
    def export_query_result(self, query_result: QueryResult, 
                          export_format: str = 'xlsx',
                          output_path: str = None) -> str:
        """Exportar resultado de consulta a formato especificado."""
        if output_path is None:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            extension = export_format.lower()
            output_path = os.path.join(config.export.output_directory, f"reporte_{timestamp}.{extension}")
        
        try:
            if export_format.lower() == 'xlsx':
                return self.create_enhanced_excel_report(query_result, output_path)
            
            elif export_format.lower() == 'csv':
                df = pd.DataFrame(query_result.preview_data, columns=query_result.columns)
                df.to_csv(output_path, index=False, encoding='utf-8-sig')
                logger.info(f"Archivo CSV creado: {output_path}")
                return output_path
            
            elif export_format.lower() == 'json':
                df = pd.DataFrame(query_result.preview_data, columns=query_result.columns)
                df.to_json(output_path, orient='records', indent=2)
                logger.info(f"Archivo JSON creado: {output_path}")
                return output_path
            
            else:
                raise ValueError(f"Formato no soportado: {export_format}")
                
        except Exception as e:
            logger.error(f"Error exportando a {export_format}", e)
            raise
    
    def get_export_statistics(self) -> Dict[str, Any]:
        """Obtener estadísticas de exportaciones."""
        active_count = len([e for e in self.active_exports.values() 
                          if isinstance(e, ExportProgress) and e.status not in ['Completado', 'Error', 'Cancelado']])
        
        completed_count = len([e for e in self.active_exports.values() 
                             if isinstance(e, ExportProgress) and e.status == 'Completado'])
        
        return {
            'total_exports': len(self.active_exports),
            'active_exports': active_count,
            'completed_exports': completed_count,
            'export_directory': config.export.output_directory
        }
    
    def cleanup_old_exports(self, max_age_hours: int = 24):
        """Limpiar exportaciones antigas."""
        cutoff_time = datetime.now().timestamp() - (max_age_hours * 3600)
        
        exports_to_remove = []
        for export_id, export_item in self.active_exports.items():
            if isinstance(export_item, ExportProgress):
                if export_item.start_time.timestamp() < cutoff_time:
                    exports_to_remove.append(export_id)
        
        for export_id in exports_to_remove:
            del self.active_exports[export_id]
        
        logger.info(f"Limpiadas {len(exports_to_remove)} exportaciones antigas")


# Instancia global del generador de reportes
report_generator = ReportGenerator()